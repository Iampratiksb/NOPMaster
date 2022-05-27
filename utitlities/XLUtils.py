import openpyxl

def getrowcount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return(sheet.max_row)

def getcolumncoubt(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_column)

def readData(file, sheetname, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum, column=columnnum).value

def writeData(file, sheetname, rownum, columnnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum, column=columnnum).value = data
    workbook.save(file)